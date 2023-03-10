from openpyxl.reader.excel import load_workbook


def get_all_user_ids() -> list:
    wb = load_workbook('db/requests.xlsx')
    ws = wb.active

    ids_column = ws['D']

    result = []
    for x in range(2, len(ids_column)):
        result.append(ids_column[x].value)

    result = list(set(result))
    if None in result:
        result.remove(None)
    print(result)

    return result
